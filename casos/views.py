# casos/views.py
# ==============================================================================
# Sistema de Gestão de Casos - Views Principais
# ==============================================================================

import logging
import re
from decimal import Decimal, InvalidOperation
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta
from django.db.models import ProtectedError

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.contrib import messages
from django.urls import reverse, reverse_lazy
from django.db.models import Sum, Q 
from django.views.decorators.http import require_POST
from django.forms import formset_factory
from django.db import transaction
from django.http import JsonResponse, HttpResponse, FileResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from io import BytesIO

# --- DRF ---
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import viewsets
from rest_framework.authentication import TokenAuthentication

# --- Imports de Outros Apps ---
from clientes.models import Cliente
from produtos.models import Produto
from campos_custom.models import (
    EstruturaDeCampos,
    CampoPersonalizado,
    GrupoCampos,
    InstanciaGrupoValor,
    ValorCampoPersonalizado,
    OpcoesListaPersonalizada
)
from integrations.sharepoint import SharePoint

# --- Imports Locais (do app 'casos') ---
from .models import (
    Caso,
    Andamento,
    ModeloAndamento,
    Timesheet,
    Acordo,
    Parcela,
    Despesa,
    FluxoInterno,
    Tomador,
    TomadorEmail,
    TomadorTelefone,
    Segurado,
    SeguradoEmail,
    SeguradoTelefone,
    Corretor,
    CorretorEmail,
    CorretorTelefone
)
from .forms import (
    CasoDinamicoForm,
    AndamentoForm,
    TimesheetForm,
    AcordoForm,
    DespesaForm,
    BaseGrupoForm,
    CasoInfoBasicasForm,
    CasoDadosAdicionaisForm,
    TomadorForm,
    SeguradoForm,
    CorretorForm
)
from .folder_utils import recriar_estrutura_de_pastas
from .serializers import CasoSerializer
import openpyxl

# PDF Imports
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

# ==============================================================================
# CONFIGURAÇÕES
# ==============================================================================
logger = logging.getLogger('casos_app')
User = get_user_model()

def normalize_currency_input(value):
    if value is None:
        return ''
    value = str(value)
    value = value.replace('R$', '').replace(' ', '').replace('\xa0', '')
    value = value.replace('.', '').replace(',', '.')
    return value

def render_titulo_caso(padrao, dados):
    if not padrao:
        return ''
    def replacer(match):
        key = match.group(1).strip().lower()
        return dados.get(key, '')
    return re.sub(r'\{([^}]+)\}', replacer, padrao)

def add_system_title_fields(dados_titulo, *, cliente=None, produto=None, caso=None, cleaned=None):
    def pick(name):
        if cleaned and name in cleaned:
            return cleaned.get(name)
        if caso is not None:
            return getattr(caso, name, None)
        return None

    def format_date(value):
        if hasattr(value, 'strftime'):
            return value.strftime('%d/%m/%Y')
        return str(value) if value else ''

    advogado = pick('advogado_responsavel')
    tomador = pick('tomador')
    segurado = pick('segurado')
    corretor = pick('corretor')
    status = pick('status')

    if cliente:
        dados_titulo.setdefault('cliente', cliente.nome)
        dados_titulo.setdefault('cliente_nome', cliente.nome)
    if produto:
        dados_titulo.setdefault('produto', produto.nome)
        dados_titulo.setdefault('produto_nome', produto.nome)

    dados_titulo.setdefault('status', status or '')
    dados_titulo.setdefault('data_entrada', format_date(pick('data_entrada')))
    dados_titulo.setdefault('data_encerramento', format_date(pick('data_encerramento')))

    if advogado:
        nome_adv = advogado.get_full_name() or advogado.username
        dados_titulo.setdefault('advogado', nome_adv)
        dados_titulo.setdefault('advogado_responsavel', nome_adv)
    if tomador:
        dados_titulo.setdefault('tomador', tomador.nome)
    if segurado:
        dados_titulo.setdefault('segurado', segurado.nome)
    if corretor:
        dados_titulo.setdefault('corretor', corretor.nome)

# Importa funções auxiliares (com fallback)
try:
    from .utils import get_cabecalho_exportacao
except ImportError:
    def get_cabecalho_exportacao(cliente=None, produto=None):
        return ([], [], {})

try:
    from .tasks import processar_linha_importacao
except ImportError:
    def processar_linha_importacao(*args, **kwargs):
        logger.critical("Tarefa Celery não encontrada!")

# ==============================================================================
# VIEWS DE TOMADOR
# ==============================================================================
@login_required
def exportar_tomadores_excel(request):
    queryset = Tomador.objects.prefetch_related('casos__cliente', 'casos__produto')
    q = request.GET.get('q')
    if q:
        queryset = queryset.filter(
            Q(nome__icontains=q) | Q(cpf__icontains=q) | Q(cnpj__icontains=q)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatorio Tomadores"
    headers = ["Nome do Tomador", "CPF/CNPJ", "Cliente", "Produto", "Número do Aviso / Título", "Status"]
    ws.append(headers)

    for tomador in queryset:
        casos = tomador.casos.all()
        if casos.exists():
            for caso in casos:
                ws.append([
                    tomador.nome, tomador.cpf or tomador.cnpj or "-",
                    caso.cliente.nome if caso.cliente else "-",
                    caso.produto.nome if caso.produto else "-",
                    caso.titulo or f"Caso #{caso.id}",
                    caso.get_status_display()
                ])
        else:
            ws.append([tomador.nome, tomador.cpf or tomador.cnpj or "-", "-", "-", "-", "Sem casos"])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_tomadores.xlsx"'
    wb.save(response)
    return response

@login_required
def exportar_segurados_excel(request):
    queryset = Segurado.objects.prefetch_related(
        'casos__cliente', 'casos__produto', 'casos__valores_personalizados__campo'
    )
    q = request.GET.get('q')
    if q:
        queryset = queryset.filter(
            Q(nome__icontains=q) | Q(cpf__icontains=q) | Q(cnpj__icontains=q)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatorio Segurados"
    headers = ["Nome do Segurado", "CPF/CNPJ", "Cliente", "Produto", "Aviso", "Status"]
    ws.append(headers)

    for segurado in queryset:
        casos = segurado.casos.all()
        if casos.exists():
            for caso in casos:
                aviso = "-"
                for valor in caso.valores_personalizados.all():
                    if valor.campo and valor.campo.nome_variavel == 'aviso' and valor.valor:
                        aviso = valor.valor
                        break
                if aviso == "-":
                    aviso = caso.titulo or f"Caso #{caso.id}"
                ws.append([
                    segurado.nome,
                    segurado.cpf or segurado.cnpj or "-",
                    caso.cliente.nome if caso.cliente else "-",
                    caso.produto.nome if caso.produto else "-",
                    aviso,
                    caso.get_status_display()
                ])
        else:
            ws.append([segurado.nome, segurado.cpf or segurado.cnpj or "-", "-", "-", "-", "Sem casos"])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_segurados.xlsx"'
    wb.save(response)
    return response

@login_required
def exportar_corretores_excel(request):
    queryset = Corretor.objects.prefetch_related(
        'casos__cliente', 'casos__produto', 'casos__valores_personalizados__campo'
    )
    q = request.GET.get('q')
    if q:
        queryset = queryset.filter(
            Q(nome__icontains=q) | Q(cpf__icontains=q) | Q(cnpj__icontains=q)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatorio Corretores"
    headers = ["Nome do Corretor", "CPF/CNPJ", "Cliente", "Produto", "Aviso", "Status"]
    ws.append(headers)

    for corretor in queryset:
        casos = corretor.casos.all()
        if casos.exists():
            for caso in casos:
                aviso = "-"
                for valor in caso.valores_personalizados.all():
                    if valor.campo and valor.campo.nome_variavel == 'aviso' and valor.valor:
                        aviso = valor.valor
                        break
                if aviso == "-":
                    aviso = caso.titulo or f"Caso #{caso.id}"
                ws.append([
                    corretor.nome,
                    corretor.cpf or corretor.cnpj or "-",
                    caso.cliente.nome if caso.cliente else "-",
                    caso.produto.nome if caso.produto else "-",
                    aviso,
                    caso.get_status_display()
                ])
        else:
            ws.append([corretor.nome, corretor.cpf or corretor.cnpj or "-", "-", "-", "-", "Sem casos"])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=\"relatorio_corretores.xlsx\"'
    wb.save(response)
    return response

@login_required
def exportar_tomadores_pdf(request):
    queryset = Tomador.objects.prefetch_related('casos__cliente', 'casos__produto')
    q = request.GET.get('q')
    if q:
        queryset = queryset.filter(
            Q(nome__icontains=q) | Q(cpf__icontains=q) | Q(cnpj__icontains=q)
        )

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_tomadores.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    style_normal = styles['Normal']
    style_normal.fontSize = 9

    elements.append(Paragraph("Relatório Detalhado de Tomadores e Casos", styles['Title']))
    elements.append(Spacer(1, 20))

    data = [["Tomador", "Cliente", "Produto", "Aviso / Título", "Status"]]
    for tomador in queryset:
        casos = tomador.casos.all()
        if casos.exists():
            for caso in casos:
                data.append([
                    Paragraph(tomador.nome[:40], style_normal),
                    Paragraph(caso.cliente.nome if caso.cliente else "-", style_normal),
                    Paragraph(caso.produto.nome if caso.produto else "-", style_normal),
                    Paragraph(str(caso.titulo)[:40], style_normal),
                    Paragraph(caso.get_status_display(), style_normal)
                ])
        else:
            data.append([Paragraph(tomador.nome, style_normal), "-", "-", "-", "Sem casos"])

    table = Table(data, colWidths=[180, 150, 150, 180, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.orange),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(table)
    doc.build(elements)
    return response

@require_POST
@login_required
def trocar_tomador_do_caso(request, pk):
    caso = get_object_or_404(Caso, pk=pk)
    novo_tomador_id = request.POST.get('novo_tomador')
    if novo_tomador_id:
        try:
            novo_tomador = Tomador.objects.get(id=novo_tomador_id)
            antigo_tomador_nome = caso.tomador.nome if caso.tomador else "Ninguém"
            caso.tomador = novo_tomador
            caso.save()
            FluxoInterno.objects.create(
                caso=caso, tipo_evento='EDICAO',
                descricao=f"Tomador alterado de '{antigo_tomador_nome}' para '{novo_tomador.nome}'.",
                autor=request.user
            )
            messages.success(request, f"✅ Tomador alterado para: {novo_tomador.nome}")
        except Tomador.DoesNotExist:
            messages.error(request, "Tomador não existe.")
    return redirect('casos:detalhe_caso', pk=pk)

class TomadorListView(ListView):
    model = Tomador
    template_name = 'casos/tomador_list.html'
    context_object_name = 'tomadores'
    paginate_by = 20
    def get_queryset(self):
        queryset = super().get_queryset()
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(nome__icontains=q) | Q(cpf__icontains=q) | Q(cnpj__icontains=q)
            )
        return queryset

class TomadorCreateView(CreateView):
    model = Tomador
    form_class = TomadorForm
    template_name = 'casos/tomador_form.html'
    success_url = reverse_lazy('casos:lista_tomadores')
    def form_valid(self, form):
        with transaction.atomic():
            self.object = form.save()
            for email in self.request.POST.getlist('lista_emails'):
                if email.strip(): TomadorEmail.objects.create(tomador=self.object, email=email.strip())
            fones = self.request.POST.getlist('lista_telefones')
            tipos = self.request.POST.getlist('lista_telefones_tipo')
            for idx, fone in enumerate(fones):
                if fone.strip():
                    tipo = tipos[idx] if idx < len(tipos) and tipos[idx] else 'CELULAR'
                    TomadorTelefone.objects.create(tomador=self.object, telefone=fone.strip(), tipo=tipo)
        return redirect(self.get_success_url())

class TomadorUpdateView(UpdateView):
    model = Tomador
    form_class = TomadorForm
    template_name = 'casos/tomador_form.html'
    success_url = reverse_lazy('casos:lista_tomadores')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['casos_vinculados'] = self.object.casos.all().select_related('cliente', 'produto')
        context['emails'] = self.object.emails.all()
        context['telefones'] = self.object.telefones.all()
        context['todos_tomadores'] = Tomador.objects.exclude(id=self.object.id).order_by('nome')
        return context
    def form_valid(self, form):
        with transaction.atomic():
            self.object = form.save()
            remove_emails = set(self.request.POST.getlist('remove_emails'))
            remove_fones = set(self.request.POST.getlist('remove_telefones'))

            for email_obj in self.object.emails.all():
                if str(email_obj.id) in remove_emails:
                    email_obj.delete()
                    continue
                new_val = self.request.POST.get(f'email_{email_obj.id}', '').strip()
                if not new_val:
                    email_obj.delete()
                elif new_val != email_obj.email:
                    email_obj.email = new_val
                    email_obj.save()

            for fone_obj in self.object.telefones.all():
                if str(fone_obj.id) in remove_fones:
                    fone_obj.delete()
                    continue
                new_val = self.request.POST.get(f'telefone_{fone_obj.id}', '').strip()
                if not new_val:
                    fone_obj.delete()
                    continue
                new_tipo = self.request.POST.get(f'telefone_tipo_{fone_obj.id}', '') or fone_obj.tipo
                changed = False
                if new_val != fone_obj.telefone:
                    fone_obj.telefone = new_val
                    changed = True
                if new_tipo != fone_obj.tipo:
                    fone_obj.tipo = new_tipo
                    changed = True
                if changed:
                    fone_obj.save()

            for email in self.request.POST.getlist('lista_emails'):
                if email.strip():
                    TomadorEmail.objects.create(tomador=self.object, email=email.strip())
            fones = self.request.POST.getlist('lista_telefones')
            tipos = self.request.POST.getlist('lista_telefones_tipo')
            for idx, fone in enumerate(fones):
                if fone.strip():
                    tipo = tipos[idx] if idx < len(tipos) and tipos[idx] else 'CELULAR'
                    TomadorTelefone.objects.create(tomador=self.object, telefone=fone.strip(), tipo=tipo)
        return redirect(self.get_success_url())

class TomadorDeleteView(DeleteView):
    model = Tomador
    template_name = 'casos/tomador_confirm_delete.html'
    success_url = reverse_lazy('casos:lista_tomadores')

    def post(self, request, *args, **kwargs):
        try:
            return super().post(request, *args, **kwargs)
        except ProtectedError:
            self.object = self.get_object()
            messages.error(
                request,
                f"Nao e possivel excluir o tomador '{self.object.nome}' pois existem casos vinculados."
            )
            return redirect('casos:editar_tomador', pk=self.object.pk)


class SeguradoListView(ListView):
    model = Segurado
    template_name = 'casos/segurado_list.html'
    context_object_name = 'segurados'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(nome__icontains=q) | Q(cpf__icontains=q) | Q(cnpj__icontains=q)
            )
        return queryset


class SeguradoCreateView(CreateView):
    model = Segurado
    form_class = SeguradoForm
    template_name = 'casos/segurado_form.html'
    success_url = reverse_lazy('casos:lista_segurados')

    def form_valid(self, form):
        with transaction.atomic():
            self.object = form.save()
            for email in self.request.POST.getlist('lista_emails'):
                if email.strip():
                    SeguradoEmail.objects.create(segurado=self.object, email=email.strip())
            fones = self.request.POST.getlist('lista_telefones')
            tipos = self.request.POST.getlist('lista_telefones_tipo')
            for idx, fone in enumerate(fones):
                if fone.strip():
                    tipo = tipos[idx] if idx < len(tipos) and tipos[idx] else 'CELULAR'
                    SeguradoTelefone.objects.create(segurado=self.object, telefone=fone.strip(), tipo=tipo)
        return redirect(self.get_success_url())


class SeguradoUpdateView(UpdateView):
    model = Segurado
    form_class = SeguradoForm
    template_name = 'casos/segurado_form.html'
    success_url = reverse_lazy('casos:lista_segurados')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['casos_vinculados'] = self.object.casos.all().select_related('cliente', 'produto')
        context['emails'] = self.object.emails.all()
        context['telefones'] = self.object.telefones.all()
        context['todos_segurados'] = Segurado.objects.exclude(id=self.object.id).order_by('nome')
        return context

    def form_valid(self, form):
        with transaction.atomic():
            self.object = form.save()
            remove_emails = set(self.request.POST.getlist('remove_emails'))
            remove_fones = set(self.request.POST.getlist('remove_telefones'))

            for email_obj in self.object.emails.all():
                if str(email_obj.id) in remove_emails:
                    email_obj.delete()
                    continue
                new_val = self.request.POST.get(f'email_{email_obj.id}', '').strip()
                if not new_val:
                    email_obj.delete()
                elif new_val != email_obj.email:
                    email_obj.email = new_val
                    email_obj.save()

            for fone_obj in self.object.telefones.all():
                if str(fone_obj.id) in remove_fones:
                    fone_obj.delete()
                    continue
                new_val = self.request.POST.get(f'telefone_{fone_obj.id}', '').strip()
                new_tipo = self.request.POST.get(f'telefone_tipo_{fone_obj.id}', fone_obj.tipo)
                if not new_val:
                    fone_obj.delete()
                elif new_val != fone_obj.telefone or new_tipo != fone_obj.tipo:
                    fone_obj.telefone = new_val
                    fone_obj.tipo = new_tipo
                    fone_obj.save()

            for email in self.request.POST.getlist('lista_emails'):
                if email.strip():
                    SeguradoEmail.objects.create(segurado=self.object, email=email.strip())
            fones = self.request.POST.getlist('lista_telefones')
            tipos = self.request.POST.getlist('lista_telefones_tipo')
            for idx, fone in enumerate(fones):
                if fone.strip():
                    tipo = tipos[idx] if idx < len(tipos) and tipos[idx] else 'CELULAR'
                    SeguradoTelefone.objects.create(segurado=self.object, telefone=fone.strip(), tipo=tipo)
        return redirect(self.get_success_url())


class SeguradoDeleteView(DeleteView):
    model = Segurado
    template_name = 'casos/segurado_confirm_delete.html'
    success_url = reverse_lazy('casos:lista_segurados')

    def post(self, request, *args, **kwargs):
        try:
            return super().post(request, *args, **kwargs)
        except ProtectedError:
            self.object = self.get_object()
            messages.error(
                request,
                f"Nao e possivel excluir o segurado '{self.object.nome}' pois existem casos vinculados."
            )
            return redirect('casos:editar_segurado', pk=self.object.pk)


class SeguradoDetailView(DetailView):
    model = Segurado
    template_name = 'casos/segurado_detail.html'
    context_object_name = 'segurado'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['casos_vinculados'] = self.object.casos.all().select_related('cliente', 'produto')
        context['emails'] = self.object.emails.all()
        context['telefones'] = self.object.telefones.all()
        return context


class CorretorListView(ListView):
    model = Corretor
    template_name = 'casos/corretor_list.html'
    context_object_name = 'corretores'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(nome__icontains=q) | Q(cpf__icontains=q) | Q(cnpj__icontains=q)
            )
        return queryset


class CorretorCreateView(CreateView):
    model = Corretor
    form_class = CorretorForm
    template_name = 'casos/corretor_form.html'
    success_url = reverse_lazy('casos:lista_corretores')

    def form_valid(self, form):
        with transaction.atomic():
            self.object = form.save()
            for email in self.request.POST.getlist('lista_emails'):
                if email.strip():
                    CorretorEmail.objects.create(corretor=self.object, email=email.strip())
            fones = self.request.POST.getlist('lista_telefones')
            tipos = self.request.POST.getlist('lista_telefones_tipo')
            for idx, fone in enumerate(fones):
                if fone.strip():
                    tipo = tipos[idx] if idx < len(tipos) and tipos[idx] else 'CELULAR'
                    CorretorTelefone.objects.create(corretor=self.object, telefone=fone.strip(), tipo=tipo)
        return redirect(self.get_success_url())


class CorretorUpdateView(UpdateView):
    model = Corretor
    form_class = CorretorForm
    template_name = 'casos/corretor_form.html'
    success_url = reverse_lazy('casos:lista_corretores')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['casos_vinculados'] = self.object.casos.all().select_related('cliente', 'produto')
        context['emails'] = self.object.emails.all()
        context['telefones'] = self.object.telefones.all()
        context['todos_corretores'] = Corretor.objects.exclude(id=self.object.id).order_by('nome')
        return context

    def form_valid(self, form):
        with transaction.atomic():
            self.object = form.save()
            remove_emails = set(self.request.POST.getlist('remove_emails'))
            remove_fones = set(self.request.POST.getlist('remove_telefones'))

            for email_obj in self.object.emails.all():
                if str(email_obj.id) in remove_emails:
                    email_obj.delete()
                    continue
                new_val = self.request.POST.get(f'email_{email_obj.id}', '').strip()
                if not new_val:
                    email_obj.delete()
                elif new_val != email_obj.email:
                    email_obj.email = new_val
                    email_obj.save()

            for fone_obj in self.object.telefones.all():
                if str(fone_obj.id) in remove_fones:
                    fone_obj.delete()
                    continue
                new_val = self.request.POST.get(f'telefone_{fone_obj.id}', '').strip()
                new_tipo = self.request.POST.get(f'telefone_tipo_{fone_obj.id}', fone_obj.tipo)
                if not new_val:
                    fone_obj.delete()
                elif new_val != fone_obj.telefone or new_tipo != fone_obj.tipo:
                    fone_obj.telefone = new_val
                    fone_obj.tipo = new_tipo
                    fone_obj.save()

            for email in self.request.POST.getlist('lista_emails'):
                if email.strip():
                    CorretorEmail.objects.create(corretor=self.object, email=email.strip())
            fones = self.request.POST.getlist('lista_telefones')
            tipos = self.request.POST.getlist('lista_telefones_tipo')
            for idx, fone in enumerate(fones):
                if fone.strip():
                    tipo = tipos[idx] if idx < len(tipos) and tipos[idx] else 'CELULAR'
                    CorretorTelefone.objects.create(corretor=self.object, telefone=fone.strip(), tipo=tipo)
        return redirect(self.get_success_url())


class CorretorDeleteView(DeleteView):
    model = Corretor
    template_name = 'casos/corretor_confirm_delete.html'
    success_url = reverse_lazy('casos:lista_corretores')

    def post(self, request, *args, **kwargs):
        try:
            return super().post(request, *args, **kwargs)
        except ProtectedError:
            self.object = self.get_object()
            messages.error(
                request,
                f"Nao e possivel excluir o corretor '{self.object.nome}' pois existem casos vinculados."
            )
            return redirect('casos:editar_corretor', pk=self.object.pk)


class CorretorDetailView(DetailView):
    model = Corretor
    template_name = 'casos/corretor_detail.html'
    context_object_name = 'corretor'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['casos_vinculados'] = self.object.casos.all().select_related('cliente', 'produto')
        context['emails'] = self.object.emails.all()
        context['telefones'] = self.object.telefones.all()
        return context

@require_POST
@login_required
def criar_tomador_ajax(request):
    form = TomadorForm(request.POST)
    if form.is_valid():
        with transaction.atomic():
            tomador = form.save()
            for email in request.POST.getlist('lista_emails'):
                if email.strip(): TomadorEmail.objects.create(tomador=tomador, email=email.strip())
            fones = request.POST.getlist('lista_telefones')
            tipos = request.POST.getlist('lista_telefones_tipo')
            for idx, fone in enumerate(fones):
                if fone.strip():
                    tipo = tipos[idx] if idx < len(tipos) and tipos[idx] else 'CELULAR'
                    TomadorTelefone.objects.create(tomador=tomador, telefone=fone.strip(), tipo=tipo)
            return JsonResponse({'success': True, 'id': tomador.id, 'text': tomador.nome})
    return JsonResponse({'success': False, 'errors': form.errors})

@require_POST
@login_required
def criar_segurado_ajax(request):
    form = SeguradoForm(request.POST)
    if form.is_valid():
        with transaction.atomic():
            segurado = form.save()
            for email in request.POST.getlist('lista_emails'):
                if email.strip():
                    SeguradoEmail.objects.create(segurado=segurado, email=email.strip())
            fones = request.POST.getlist('lista_telefones')
            tipos = request.POST.getlist('lista_telefones_tipo')
            for idx, fone in enumerate(fones):
                if fone.strip():
                    tipo = tipos[idx] if idx < len(tipos) and tipos[idx] else 'CELULAR'
                    SeguradoTelefone.objects.create(segurado=segurado, telefone=fone.strip(), tipo=tipo)
            return JsonResponse({'success': True, 'id': segurado.id, 'text': segurado.nome})
    return JsonResponse({'success': False, 'errors': form.errors})

@require_POST
@login_required
def criar_corretor_ajax(request):
    form = CorretorForm(request.POST)
    if form.is_valid():
        with transaction.atomic():
            corretor = form.save()
            for email in request.POST.getlist('lista_emails'):
                if email.strip():
                    CorretorEmail.objects.create(corretor=corretor, email=email.strip())
            fones = request.POST.getlist('lista_telefones')
            tipos = request.POST.getlist('lista_telefones_tipo')
            for idx, fone in enumerate(fones):
                if fone.strip():
                    tipo = tipos[idx] if idx < len(tipos) and tipos[idx] else 'CELULAR'
                    CorretorTelefone.objects.create(corretor=corretor, telefone=fone.strip(), tipo=tipo)
            return JsonResponse({'success': True, 'id': corretor.id, 'text': corretor.nome})
    return JsonResponse({'success': False, 'errors': form.errors})

# ==============================================================================
# CRIAR CASO (LOGICA DE MOEDA ATUALIZADA)
# ==============================================================================
@login_required
def criar_caso(request, cliente_id, produto_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    produto = get_object_or_404(Produto, id=produto_id)

    try:
        estrutura = EstruturaDeCampos.objects.prefetch_related(
            'ordenamentos_simples__campo', 'grupos_repetiveis__ordenamentos_grupo__campo'
        ).get(cliente=cliente, produto=produto)
    except EstruturaDeCampos.DoesNotExist:
        messages.error(request, "Estrutura não definida.")
        return redirect('casos:selecionar_produto_cliente')

    if request.method == 'POST':
        post_data = request.POST.copy()
        if post_data.get('valor_apurado'):
            post_data['valor_apurado'] = normalize_currency_input(post_data['valor_apurado'])

        for key, value in post_data.items():
            if 'campo_personalizado' in key and value and isinstance(value, str):
                if ',' in value or 'R$' in value:
                    post_data[key] = normalize_currency_input(value)

        form = CasoDinamicoForm(post_data, cliente=cliente, produto=produto)
    else:
        form = CasoDinamicoForm(cliente=cliente, produto=produto)

    grupo_formsets = {}
    for grupo in estrutura.grupos_repetiveis.all():
        GrupoFormSet = formset_factory(BaseGrupoForm, extra=0, can_delete=True)
        prefix = f'grupo_{grupo.id}'
        kwargs = {'grupo_campos': grupo, 'cliente': cliente, 'produto': produto}
        if request.method == 'POST':
            formset = GrupoFormSet(post_data, prefix=prefix, form_kwargs=kwargs)
        else:
            formset = GrupoFormSet(prefix=prefix, form_kwargs=kwargs)
        grupo_formsets[grupo.id] = (grupo, formset)

    if request.method == 'POST':
        formsets_validos = all(fs.is_valid() for _, fs in grupo_formsets.values())
        if form.is_valid() and formsets_validos:
            try:
                with transaction.atomic():
                    # Lógica de Título Automático
                    dados_titulo = {}
                    dados_limpos = form.cleaned_data
                    for eco in estrutura.ordenamentos_simples.all():
                        val = dados_limpos.get(f'campo_personalizado_{eco.campo.id}', '')
                        dados_titulo[eco.campo.nome_variavel.lower()] = '' if val is None else str(val)

                    for grupo, formset in grupo_formsets.values():
                        for f in formset:
                            if not f.has_changed() or f.cleaned_data.get('DELETE'):
                                continue
                            for conf in grupo.ordenamentos_grupo.all():
                                val_f = f.cleaned_data.get(f'campo_personalizado_{conf.campo.id}')
                                if val_f not in (None, ''):
                                    dados_titulo.setdefault(
                                        conf.campo.nome_variavel.lower(),
                                        str(val_f)
                                    )

                    add_system_title_fields(dados_titulo, cliente=cliente, produto=produto, cleaned=dados_limpos)

                    if produto.padrao_titulo:
                        titulo_final = render_titulo_caso(produto.padrao_titulo or "", dados_titulo)
                        if not titulo_final.strip():
                            titulo_final = f"Caso {cliente.nome}"
                    else:
                        titulo_manual = dados_limpos.get('titulo_manual', '').strip()
                        titulo_final = titulo_manual if titulo_manual else f"Caso {cliente.nome}"

                    novo_caso = form.save(commit=False)
                    novo_caso.cliente = cliente
                    novo_caso.produto = produto
                    novo_caso.titulo = titulo_final
                    novo_caso._criador = request.user
                    novo_caso.save()

                    # Salvar campos simples
                    for eco in estrutura.ordenamentos_simples.all():
                        val = dados_limpos.get(f'campo_personalizado_{eco.campo.id}')
                        if val is not None:
                            ValorCampoPersonalizado.objects.create(caso=novo_caso, campo=eco.campo, valor=str(val))

                    # Salvar formsets
                    for grupo, formset in grupo_formsets.values():
                        for idx, f in enumerate(formset):
                            if not f.has_changed() or f.cleaned_data.get('DELETE'): continue
                            inst = InstanciaGrupoValor.objects.create(caso=novo_caso, grupo=grupo, ordem_instancia=idx)
                            for conf in grupo.ordenamentos_grupo.all():
                                val_f = f.cleaned_data.get(f'campo_personalizado_{conf.campo.id}')
                                ValorCampoPersonalizado.objects.create(instancia_grupo=inst, campo=conf.campo, valor=str(val_f))

                    messages.success(request, f"✅ Caso criado!")
                    return redirect('casos:detalhe_caso', pk=novo_caso.pk)
            except Exception as e:
                messages.error(request, f"Erro: {e}")

    return render(request, 'casos/criar_caso_form.html', {
        'cliente': cliente, 'produto': produto, 'form': form,
        'grupo_formsets': grupo_formsets.values(), 'estrutura': estrutura
    })


# ==============================================================================
# EDITAR CASO
# ==============================================================================
@login_required
def editar_caso(request, pk):
    caso = get_object_or_404(Caso, pk=pk)
    cliente = caso.cliente
    produto = caso.produto

    try:
        estrutura = EstruturaDeCampos.objects.prefetch_related(
            'campos', 'grupos_repetiveis__campos'
        ).get(cliente=cliente, produto=produto)
    except EstruturaDeCampos.DoesNotExist:
        messages.error(request, "Estrutura de campos não definida.")
        return redirect('casos:lista_casos')

    if request.method == 'POST':
        post_data = request.POST.copy()
        if post_data.get('valor_apurado'):
            post_data['valor_apurado'] = normalize_currency_input(post_data['valor_apurado'])

        for key, value in post_data.items():
            if 'campo_personalizado' in key and value and isinstance(value, str):
                if ',' in value or 'R$' in value:
                    post_data[key] = normalize_currency_input(value)

        form = CasoDinamicoForm(post_data, instance=caso, cliente=cliente, produto=produto)
    else:
        form = CasoDinamicoForm(instance=caso, cliente=cliente, produto=produto)

    grupo_formsets = {}
    for grupo in estrutura.grupos_repetiveis.all():
        GrupoFormSet = formset_factory(BaseGrupoForm, extra=0, can_delete=True)
        prefix = f'grupo_{grupo.id}'
        kwargs = {'grupo_campos': grupo, 'cliente': cliente, 'produto': produto}

        instancias_salvas = caso.grupos_de_valores.filter(grupo=grupo).prefetch_related('valores__campo')
        initial_data = []
        for instancia in instancias_salvas:
            dados_instancia = {}
            for valor in instancia.valores.all():
                dados_instancia[f'campo_personalizado_{valor.campo.id}'] = valor.valor
            initial_data.append(dados_instancia)

        if request.method == 'POST':
            formset = GrupoFormSet(post_data, prefix=prefix, form_kwargs=kwargs, initial=initial_data)
        else:
            formset = GrupoFormSet(prefix=prefix, form_kwargs=kwargs, initial=initial_data)
        grupo_formsets[grupo.id] = (grupo, formset)

    if request.method == 'POST':
        formsets_validos = all(fs.is_valid() for _, fs in grupo_formsets.values())
        if form.is_valid() and formsets_validos:
            try:
                with transaction.atomic():
                    caso = form.save(commit=False)

                    dados_limpos = form.cleaned_data
                    dados_titulo = {}
                    for eco in estrutura.ordenamentos_simples.all():
                        val = dados_limpos.get(f'campo_personalizado_{eco.campo.id}', '')
                        dados_titulo[eco.campo.nome_variavel.lower()] = '' if val is None else str(val)

                    for grupo, formset in grupo_formsets.values():
                        for f in formset:
                            if not f.has_changed() or f.cleaned_data.get('DELETE'):
                                continue
                            for conf in grupo.ordenamentos_grupo.all():
                                val_f = f.cleaned_data.get(f'campo_personalizado_{conf.campo.id}')
                                if val_f not in (None, ''):
                                    dados_titulo.setdefault(
                                        conf.campo.nome_variavel.lower(),
                                        str(val_f)
                                    )

                    add_system_title_fields(dados_titulo, cliente=cliente, produto=produto, caso=caso, cleaned=dados_limpos)
                    if produto.padrao_titulo:
                        titulo_final = render_titulo_caso(produto.padrao_titulo or "", dados_titulo)
                        if not titulo_final.strip():
                            titulo_final = f"Caso {cliente.nome}"
                        caso.titulo = titulo_final
                    else:
                        titulo_manual = dados_limpos.get('titulo_manual', '').strip()
                        if titulo_manual:
                            caso.titulo = titulo_manual

                    caso.save()

                    for eco in estrutura.ordenamentos_simples.all():
                        val = form.cleaned_data.get(f'campo_personalizado_{eco.campo.id}')
                        valor_obj, _ = ValorCampoPersonalizado.objects.get_or_create(
                            caso=caso, campo=eco.campo, instancia_grupo__isnull=True
                        )
                        valor_obj.valor = '' if val is None else str(val)
                        valor_obj.save()

                    for grupo, formset in grupo_formsets.values():
                        caso.grupos_de_valores.filter(grupo=grupo).delete()
                        for idx, f in enumerate(formset):
                            if not f.has_changed() or f.cleaned_data.get('DELETE'):
                                continue
                            inst = InstanciaGrupoValor.objects.create(
                                caso=caso, grupo=grupo, ordem_instancia=idx
                            )
                            for conf in grupo.ordenamentos_grupo.all():
                                val_f = f.cleaned_data.get(f'campo_personalizado_{conf.campo.id}')
                                ValorCampoPersonalizado.objects.create(
                                    instancia_grupo=inst, campo=conf.campo, valor=str(val_f)
                                )

                    messages.success(request, "Caso atualizado!")
                    return redirect('casos:detalhe_caso', pk=caso.pk)
            except Exception as e:
                messages.error(request, f"Erro: {e}")

    return render(request, 'casos/editar_caso.html', {
        'cliente': cliente,
        'produto': produto,
        'caso': caso,
        'form': form,
        'grupo_formsets': grupo_formsets.values(),
        'estrutura': estrutura
    })


@login_required
def lista_casos(request):
    casos_list = Caso.objects.select_related(
        'cliente', 'produto', 'advogado_responsavel'
    ).all().order_by('-id')
    
    filtro_titulo = request.GET.get('filtro_titulo', '')
    filtro_cliente = request.GET.get('filtro_cliente', '')
    filtro_produto = request.GET.get('filtro_produto', '')
    filtro_status = request.GET.get('filtro_status', '')
    filtro_advogado = request.GET.get('filtro_advogado', '')

    if filtro_titulo:
        casos_list = casos_list.filter(titulo__icontains=filtro_titulo)
    if filtro_cliente:
        casos_list = casos_list.filter(cliente_id=filtro_cliente)
    if filtro_produto:
        casos_list = casos_list.filter(produto_id=filtro_produto)
    if filtro_status:
        casos_list = casos_list.filter(status=filtro_status)
    if filtro_advogado:
        casos_list = casos_list.filter(advogado_responsavel_id=filtro_advogado)

    paginator = Paginator(casos_list, 20)
    page = request.GET.get('page')
    casos = paginator.get_page(page)

    context = {
        'casos': casos,
        'valores_filtro': request.GET,
        'todos_clientes': Cliente.objects.all().order_by('nome'),
        'todos_produtos': Produto.objects.all().order_by('nome'),
        'todos_advogados': User.objects.all().order_by('first_name', 'username'),
        'status_choices': Caso.STATUS_CHOICES,
    }
    return render(request, 'casos/lista_casos.html', context)


# ==============================================================================
# DETALHE DO CASO
# ==============================================================================
@login_required
def detalhe_caso(request, pk):
    caso = get_object_or_404(Caso, pk=pk)
    form_acordo = AcordoForm(user=request.user)

    if request.method == 'POST':
        if 'submit_despesa' in request.POST:
            data = request.POST.copy()
            valor = normalize_currency_input(data.get('valor', ''))
            if valor:
                data['valor'] = valor
            form_despesa = DespesaForm(data, request.FILES, user=request.user)
            if form_despesa.is_valid():
                despesa = form_despesa.save(commit=False)
                despesa.caso = caso
                despesa.save()
                messages.success(request, 'Despesa registrada com sucesso.')
                return redirect(f"{reverse('casos:detalhe_caso', kwargs={'pk': caso.pk})}?aba=despesas")
            messages.error(request, 'Corrija os erros da despesa.')

        if 'submit_despesa_edit' in request.POST:
            despesa_id = request.POST.get('despesa_id')
            despesa = get_object_or_404(Despesa, pk=despesa_id, caso=caso)
            data = request.POST.copy()
            valor = normalize_currency_input(data.get('valor', ''))
            if valor:
                data['valor'] = valor
            form_edit = DespesaForm(data, request.FILES, instance=despesa, user=request.user)
            if form_edit.is_valid():
                form_edit.save()
                messages.success(request, 'Despesa atualizada com sucesso.')
                return redirect(f"{reverse('casos:detalhe_caso', kwargs={'pk': caso.pk})}?aba=despesas")
            messages.error(request, 'Corrija os erros da despesa.')

        if 'submit_timesheet' in request.POST:
            form_ts = TimesheetForm(request.POST, user=request.user)
            if form_ts.is_valid():
                ts = form_ts.save(commit=False)
                ts.caso = caso
                ts.save()
                messages.success(request, 'Timesheet registrado com sucesso.')
                return redirect(f"{reverse('casos:detalhe_caso', kwargs={'pk': caso.pk})}?aba=timesheet")
            messages.error(request, 'Corrija os erros do timesheet.')

        if 'submit_timesheet_edit' in request.POST:
            ts_id = request.POST.get('timesheet_id')
            ts_obj = get_object_or_404(Timesheet, pk=ts_id, caso=caso)
            form_edit = TimesheetForm(request.POST, instance=ts_obj, user=request.user)
            if form_edit.is_valid():
                form_edit.save()
                messages.success(request, 'Timesheet atualizado com sucesso.')
                return redirect(f"{reverse('casos:detalhe_caso', kwargs={'pk': caso.pk})}?aba=timesheet")
            messages.error(request, 'Corrija os erros do timesheet.')

        if 'submit_acordo' in request.POST:
            data = request.POST.copy()
            valor_total = normalize_currency_input(data.get('valor_total', ''))
            if valor_total:
                data['valor_total'] = valor_total
            form_acordo = AcordoForm(data, user=request.user)
            if form_acordo.is_valid():
                with transaction.atomic():
                    acordo = form_acordo.save(commit=False)
                    acordo.caso = caso
                    acordo.save()

                    valor_total_calc = acordo.valor_total
                    num_parcelas = acordo.numero_parcelas
                    if num_parcelas and valor_total_calc is not None:
                        valor_parcela = (valor_total_calc / num_parcelas).quantize(Decimal('0.01'))
                        for i in range(num_parcelas):
                            data_vencimento = acordo.data_primeira_parcela + relativedelta(months=i)
                            Parcela.objects.create(
                                acordo=acordo,
                                numero_parcela=i + 1,
                                valor_parcela=valor_parcela,
                                data_vencimento=data_vencimento
                            )
                        total_calculado = valor_parcela * num_parcelas
                        diff = valor_total_calc - total_calculado
                        if diff:
                            ultima_parcela = acordo.parcelas.order_by('-numero_parcela').first()
                            if ultima_parcela:
                                ultima_parcela.valor_parcela += diff
                                ultima_parcela.save()

                messages.success(request, 'Acordo criado com sucesso.')
                return redirect(f"{reverse('casos:detalhe_caso', kwargs={'pk': caso.pk})}?aba=acordos")
            else:
                messages.error(request, 'Corrija os erros do acordo.')

        edit_modal = request.POST.get('edit_modal')
        edit_modal = request.POST.get('edit_modal')
        
        if edit_modal == 'info-basicas':
            try:
                caso.status = request.POST.get('status', caso.status)
                data_entrada = request.POST.get('data_entrada')
                if data_entrada:
                    caso.data_entrada = data_entrada
                
                valor_apurado = normalize_currency_input(request.POST.get('valor_apurado', ''))
                if valor_apurado:
                    caso.valor_apurado = Decimal(valor_apurado)
                
                advogado_id = request.POST.get('advogado_responsavel')
                if advogado_id:
                    caso.advogado_responsavel_id = advogado_id
                else:
                    caso.advogado_responsavel = None
                
                caso.save()
                FluxoInterno.objects.create(
                    caso=caso, tipo_evento='EDICAO', descricao='Informações básicas do caso foram atualizadas.', autor=request.user
                )
                messages.success(request, '✅ Informações básicas atualizadas com sucesso!')
                return redirect('casos:detalhe_caso', pk=caso.pk)
            except Exception as e:
                messages.error(request, f'❌ Erro ao atualizar: {str(e)}')
                return redirect('casos:detalhe_caso', pk=caso.pk)
        
        elif edit_modal == 'dados-adicionais':
            # Atualiza campos personalizados dinâmicos
            for key, value in request.POST.items():
                if key.startswith('campo_'):
                    c_id = key.replace('campo_', '')
                    # Limpeza rápida se for moeda (detectado por vírgula)
                    if isinstance(value, str) and (',' in value or 'R$' in value):
                        value = normalize_currency_input(value)
                    
                    val_obj, _ = ValorCampoPersonalizado.objects.get_or_create(
                        caso=caso, campo_id=c_id, instancia_grupo__isnull=True
                    )
                    val_obj.valor = value
                    val_obj.save()
            messages.success(request, '✅ Dados atualizados!')
            return redirect('casos:detalhe_caso', pk=pk)

    # Renderização normal do detalhe
    estrutura = EstruturaDeCampos.objects.filter(
        cliente=caso.cliente, produto=caso.produto
    ).prefetch_related('grupos_repetiveis__ordenamentos_grupo__campo').first()

    grupos_repetiveis_context = []
    if estrutura:
        for grupo in estrutura.grupos_repetiveis.all():
            campos_grupo = list(grupo.ordenamentos_grupo.select_related('campo').order_by('order'))
            instancias = caso.grupos_de_valores.filter(grupo=grupo).prefetch_related('valores__campo')
            instancias_context = []
            for instancia in instancias:
                valores_map = {v.campo_id: v.valor for v in instancia.valores.all()}
                valores_render = []
                for conf in campos_grupo:
                    valores_render.append({
                        'campo': conf.campo,
                        'valor': valores_map.get(conf.campo_id, '')
                    })
                instancias_context.append({'valores': valores_render, 'placeholder': False})
            if not instancias_context:
                valores_render = [{'campo': conf.campo, 'valor': ''} for conf in campos_grupo]
                instancias_context.append({'valores': valores_render, 'placeholder': True})
            grupos_repetiveis_context.append({'grupo': grupo, 'instancias': instancias_context})

        saldo_devedor_total = Parcela.objects.filter(acordo__caso=caso, status='EMITIDA').aggregate(total=Sum('valor_parcela'))['total']

    context = {
        'caso': caso,
        'form_andamento': AndamentoForm(),
        'form_timesheet': TimesheetForm(user=request.user),
        'timesheet_forms': {t.pk: TimesheetForm(instance=t, user=request.user) for t in caso.timesheets.all()},
        'form_acordo': form_acordo,
        'form_despesa': DespesaForm(user=request.user),
        'despesa_forms': {d.pk: DespesaForm(instance=d, user=request.user) for d in caso.despesas.all()},
        'andamentos': caso.andamentos.all().order_by('-data_andamento'),
        'timesheets': caso.timesheets.all(),
        'acordos': caso.acordos.all(),
        'despesas': caso.despesas.all(),
        'fluxo_interno': caso.fluxo_interno.all().order_by('-data_evento'),
        'form_info_basicas': CasoInfoBasicasForm(instance=caso),
        'valores_personalizados': caso.valores_personalizados.filter(instancia_grupo__isnull=True).select_related('campo'),
        'grupos_de_valores_salvos': caso.grupos_de_valores.all().prefetch_related('valores__campo'),
        'grupos_repetiveis_context': grupos_repetiveis_context,
        'saldo_devedor_total': saldo_devedor_total,
    }
    return render(request, 'casos/detalhe_caso.html', context)

# ==============================================================================
# DASHBOARD EXECUTIVO
# ==============================================================================
@login_required
def dashboard_view(request):
    ano_atual = timezone.now().year
    anos_disponiveis = list(range(ano_atual, ano_atual - 5, -1))
    ano_selecionado = request.GET.get('ano', str(ano_atual))
    
    casos_ano = Caso.objects.filter(data_entrada__year=ano_selecionado)
    
    context = {
        'ano_selecionado': ano_selecionado,
        'anos_disponiveis': anos_disponiveis,
        'total_casos_ano': casos_ano.count(),
        'casos_ativos': casos_ano.filter(status='ATIVO').count(),
        'periodo': request.GET.get('periodo', 'hoje'),
    }
    return render(request, 'casos/dashboard.html', context)

# ==============================================================================
# EXPORTAR ANDAMENTOS
# ==============================================================================
@login_required
def exportar_andamentos_excel(request, pk):
    caso = get_object_or_404(Caso, pk=pk)
    andamentos = caso.andamentos.select_related('autor').order_by('data_andamento')
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f'Andamentos Caso #{caso.id}'
    headers = ['Data do Andamento', 'Descrição', 'Criado por', 'Data de Criação']
    sheet.append(headers)
    for andamento in andamentos:
        autor_nome = '-'
        if andamento.autor:
            autor_nome = andamento.autor.get_full_name() or andamento.autor.username
        data_andamento_formatada = andamento.data_andamento.strftime('%d/%m/%Y')
        data_criacao_formatada = timezone.localtime(andamento.data_criacao).strftime('%d/%m/%Y %H:%M')
        sheet.append([data_andamento_formatada, andamento.descricao, autor_nome, data_criacao_formatada])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="andamentos_caso_{caso.id}.xlsx"'
    workbook.save(response)
    return response


# ==============================================================================
# SHAREPOINT & ANEXOS
# ==============================================================================
@login_required
def carregar_painel_anexos(request, pk):
    caso = get_object_or_404(Caso, pk=pk)
    modo = request.GET.get('modo', 'anexos')

    if not caso.sharepoint_folder_id:
        return render(request, 'casos/partials/painel_anexos_criar.html', {'caso': caso})

    try:
        sp = SharePoint()
        itens = sp.listar_conteudo_pasta(caso.sharepoint_folder_id)
        context = {
            'caso': caso,
            'itens': itens,
            'folder_id': caso.sharepoint_folder_id,
            'root_folder_id': caso.sharepoint_folder_id,
            'folder_name': f"Caso #{caso.id}",
            'modo': modo
        }
        if modo == 'analyser':
            return render(request, 'casos/partials/painel_anexos_analyser.html', context)
        return render(request, 'casos/partials/painel_anexos.html', context)
    except Exception as e:
        logger.error(f"Erro ao carregar anexos: {e}", exc_info=True)
        return render(request, 'casos/partials/painel_anexos_erro.html', {
            'caso': caso,
            'mensagem_erro': f"Erro ao conectar ao SharePoint: {str(e)}"
        })

@login_required
def carregar_painel_analyser(request, pk):
    request.GET = request.GET.copy()
    request.GET._mutable = True
    request.GET['modo'] = 'analyser'
    return carregar_painel_anexos(request, pk)

@require_POST
@login_required
def criar_pasta_para_caso(request, pk):
    caso = get_object_or_404(Caso, pk=pk)
    try:
        sp = SharePoint()
        nome_pasta_caso = str(caso.id)
        pasta_caso_id = sp.criar_pasta_caso(nome_pasta_caso)
        caso.sharepoint_folder_id = pasta_caso_id
        caso.save()
        return carregar_painel_anexos(request, pk)
    except Exception as e:
        logger.error(f"Erro ao criar pasta: {e}", exc_info=True)
        return render(request, 'casos/partials/painel_anexos_erro.html', {
            'caso': caso,
            'mensagem_erro': f"Erro ao criar pasta: {e}"
        })

@login_required
def recriar_pastas_sharepoint(request, pk):
    caso = get_object_or_404(Caso, pk=pk)
    try:
        caso.sharepoint_folder_id = None
        caso.save(update_fields=['sharepoint_folder_id'])
        folder_id = recriar_estrutura_de_pastas(caso)
        sp = SharePoint()
        conteudo = sp.listar_conteudo_pasta(folder_id)
        context = {
            'caso': caso,
            'itens': conteudo,
            'folder_id': folder_id,
            'root_folder_id': folder_id,
            'folder_name': "Raiz"
        }
        return render(request, 'casos/partials/lista_arquivos.html', context)
    except Exception as e:
        return HttpResponse(f"<div class='alert alert-danger'><strong>Falha ao recriar pastas:</strong> {e}</div>")

@login_required
def baixar_arquivo_sharepoint(request, caso_pk, arquivo_id):
    caso = get_object_or_404(Caso, pk=caso_pk)
    try:
        sp = SharePoint()
        conteudo = sp.baixar_arquivo(arquivo_id)
        info = sp.obter_info_arquivo(arquivo_id)
        nome = info.get('name', 'arquivo')

        response = HttpResponse(conteudo, content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{nome}"'
        return response
    except Exception as e:
        logger.error(f"Erro no download: {e}", exc_info=True)
        messages.error(request, f"Erro ao baixar: {str(e)}")
        return redirect('casos:detalhe_caso', pk=caso_pk)

@login_required
def deletar_arquivo_sharepoint(request, caso_pk):
    caso = get_object_or_404(Caso, pk=caso_pk)
    arquivo_id = request.GET.get('arquivo_id')

    if not arquivo_id:
        return JsonResponse({'error': 'ID do arquivo nao fornecido'}, status=400)

    try:
        sp = SharePoint()
        sp.excluir_item(arquivo_id)
        return carregar_painel_anexos(request, caso_pk)
    except Exception as e:
        logger.error(f"Erro ao deletar arquivo: {e}", exc_info=True)
        return render(request, 'casos/partials/painel_anexos_erro.html', {
            'caso': caso,
            'mensagem_erro': f"Erro ao deletar arquivo: {str(e)}"
        })

@login_required
def criar_pasta_sharepoint(request, caso_pk):
    caso = get_object_or_404(Caso, pk=caso_pk)

    if request.method != 'POST':
        return JsonResponse({'error': 'Metodo nao permitido'}, status=405)

    try:
        nome_pasta = request.POST.get('nome_pasta', '').strip()
        if not nome_pasta:
            return JsonResponse({'error': 'Nome obrigatorio'}, status=400)

        sp = SharePoint()
        sp.criar_subpasta(caso.sharepoint_folder_id, nome_pasta)
        return carregar_painel_anexos(request, caso_pk)
    except Exception as e:
        logger.error(f"Erro ao criar subpasta: {e}", exc_info=True)
        return render(request, 'casos/partials/painel_anexos_erro.html', {
            'caso': caso,
            'mensagem_erro': f"Erro: {str(e)}"
        })

@login_required
def carregar_conteudo_pasta(request, folder_id):
    caso_pk = request.GET.get('caso_pk')
    root_folder_id = request.GET.get('root_folder_id', folder_id)
    modo = request.GET.get('modo', 'anexos')

    caso = get_object_or_404(Caso, pk=caso_pk) if caso_pk else None

    try:
        sp = SharePoint()
        folder_details = sp.get_item_details(folder_id)
        itens = sp.listar_conteudo_pasta(folder_id)

        context = {
            'caso': caso,
            'itens': itens,
            'folder_id': folder_id,
            'root_folder_id': root_folder_id,
            'folder_name': folder_details.get('name', 'Pasta'),
            'folder_details': folder_details,
            'modo': modo
        }

        if modo == 'analyser':
            return render(request, 'casos/partials/painel_anexos_analyser.html', context)
        return render(request, 'casos/partials/painel_anexos.html', context)
    except Exception as e:
        logger.error(f"Erro ao carregar pasta: {e}", exc_info=True)
        return HttpResponse(f"<div class='alert alert-danger'>Erro: {e}</div>")

@login_required
def preview_anexo(request, item_id):
    try:
        sp = SharePoint()
        preview_url = sp.get_preview_url(item_id)
        return HttpResponse(f'<iframe src="{preview_url}"></iframe>')
    except Exception as e:
        return HttpResponse(f"<p style='color:red;'>Erro: {e}</p>")

@require_POST
@login_required
def excluir_anexo_sharepoint(request, item_id):
    try:
        sp = SharePoint()
        sp.excluir_item(item_id)
        response = HttpResponse(status=200)
        response['HX-Refresh'] = 'true'
        return response
    except Exception as e:
        return HttpResponse(f"<p style='color:red;'>Erro: {e}</p>", status=400)

@login_required
def listar_arquivos_para_analise(request, pk):
    caso = get_object_or_404(Caso, pk=pk)
    try:
        sp = SharePoint()
        if not caso.sharepoint_folder_id:
            return JsonResponse({'success': False, 'arquivos': [], 'mensagem': 'Pasta nao encontrada'})

        itens = sp.listar_conteudo_pasta(caso.sharepoint_folder_id)
        arquivos = [
            {'id': item['id'], 'nome': item.get('name') or item.get('nome')}
            for item in itens
            if item.get('type') == 'file' or item.get('tipo') == 'file'
        ]
        return JsonResponse({'success': True, 'arquivos': arquivos})
    except Exception as e:
        return JsonResponse({'success': False, 'arquivos': [], 'mensagem': str(e)})

@login_required
def analyser_navegador_pasta(request, pk, folder_id):
    caso = get_object_or_404(Caso, pk=pk)
    root_folder_id = request.GET.get('root_folder_id', folder_id)
    try:
        sp = SharePoint()
        folder_details = sp.get_item_details(folder_id)
        itens = sp.listar_conteudo_pasta(folder_id)
        context = {
            'caso': caso,
            'itens': itens,
            'folder_id': folder_id,
            'root_folder_id': root_folder_id,
            'folder_name': folder_details.get('name', 'Pasta'),
            'folder_details': folder_details
        }
        return render(request, 'casos/partials/analyser_navegador.html', context)
    except Exception as e:
        return HttpResponse(f'<div class="alert alert-danger">Erro: {e}</div>')

@login_required
def exportar_timesheet_excel(request, pk):
    caso = get_object_or_404(Caso, pk=pk)
    timesheets = caso.timesheets.select_related('advogado').order_by('data_execucao')
    soma_total_obj = timesheets.aggregate(total_tempo=Sum('tempo'))
    tempo_total = soma_total_obj['total_tempo'] or timedelta(0)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f'Timesheet Caso #{caso.id}'
    headers = ['Data da Execução', 'Advogado', 'Descrição', 'Tempo Gasto']
    sheet.append(headers)
    for ts in timesheets:
        advogado_nome = ts.advogado.get_full_name() or ts.advogado.username if ts.advogado else '-'
        tempo_str = str(ts.tempo)
        sheet.append([ts.data_execucao.strftime('%d/%m/%Y'), advogado_nome, ts.descricao, tempo_str])
    sheet.append([])
    from openpyxl.styles import Font
    bold_font = Font(bold=True)
    linha_total = ['', '', 'Total:', str(tempo_total)]
    sheet.append(linha_total)
    sheet['C' + str(sheet.max_row)].font = bold_font
    sheet['D' + str(sheet.max_row)].font = bold_font
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="timesheet_caso_{caso.id}.xlsx"'
    workbook.save(response)
    return response


@login_required
def exportar_timesheet_pdf(request, pk):
    caso = get_object_or_404(Caso, pk=pk)
    timesheets = caso.timesheets.select_related('advogado').order_by('data_execucao')
    soma_total_obj = timesheets.aggregate(total_tempo=Sum('tempo'))
    tempo_total = soma_total_obj['total_tempo'] or timedelta(0)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=inch/2, leftMargin=inch/2, topMargin=inch/2, bottomMargin=inch/2)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph(f"Relatório de Timesheet - Caso #{caso.id}", styles['h1']))
    story.append(Paragraph(f"<b>Cliente:</b> {caso.cliente.nome}", styles['Normal']))
    story.append(Paragraph(f"<b>Produto:</b> {caso.produto.nome}", styles['Normal']))
    story.append(Spacer(1, 0.25*inch))
    data = [['Data', 'Advogado', 'Descrição', 'Tempo Gasto']]
    for ts in timesheets:
        advogado_nome = ts.advogado.get_full_name() or ts.advogado.username if ts.advogado else '-'
        data.append([
            ts.data_execucao.strftime('%d/%m/%Y'),
            advogado_nome,
            Paragraph(ts.descricao.replace('\n', '<br/>'), styles['Normal']),
            str(ts.tempo)
        ])
    data.append(['', '', Paragraph("<b>Total:</b>", styles['Normal']), Paragraph(f"<b>{str(tempo_total)}</b>", styles['Normal'])])
    table = Table(data, colWidths=[1.2*inch, 1.5*inch, 3.3*inch, 1.2*inch])
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
        ('ALIGN', (2, -1), (3, -1), 'RIGHT'),
        ('FONTNAME', (-2, -1), (-1, -1), 'Helvetica-Bold'),
    ])
    table.setStyle(style)
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="timesheet_caso_{caso.id}.pdf"'
    return response


# ==============================================================================
# EDIÇÃO DE TIMESHEET/DESPESA/ACORDO
# ==============================================================================

@login_required
def editar_timesheet(request, pk):
    timesheet = get_object_or_404(Timesheet, pk=pk)
    caso = timesheet.caso
    if request.method == 'POST':
        form = TimesheetForm(request.POST, instance=timesheet)
        if form.is_valid():
            ts_editado = form.save(commit=False)
            tempo_str = request.POST.get('tempo')
            try:
                horas, minutos = map(int, tempo_str.split(':'))
                ts_editado.tempo = timedelta(hours=horas, minutes=minutos)
            except (ValueError, TypeError):
                form.add_error('tempo', 'Formato inválido. Use HH:MM.')
            else:
                ts_editado.save()
                return redirect(f"{reverse('casos:detalhe_caso', kwargs={'pk': caso.pk})}?aba=timesheet")
    else:
        initial_data = {}
        if timesheet.tempo:
            total_seconds = int(timesheet.tempo.total_seconds())
            horas = total_seconds // 3600
            minutos = (total_seconds % 3600) // 60
            initial_data['tempo'] = f"{str(horas).zfill(2)}:{str(minutos).zfill(2)}"
        form = TimesheetForm(instance=timesheet, initial=initial_data)
    context = {'form': form, 'timesheet': timesheet, 'caso': caso}
    return render(request, 'casos/timesheet_form.html', context)


@login_required
def deletar_timesheet(request, pk):
    timesheet = get_object_or_404(Timesheet, pk=pk)
    caso = timesheet.caso
    if request.method == 'POST':
        timesheet.delete()
        return redirect(f"{reverse('casos:detalhe_caso', kwargs={'pk': caso.pk})}?aba=timesheet")
    context = {'timesheet': timesheet, 'caso': caso}
    return render(request, 'casos/timesheet_confirm_delete.html', context)


@require_POST
@login_required
def quitar_parcela(request, pk):
    parcela = get_object_or_404(Parcela, pk=pk)
    if parcela.status == 'QUITADA':
        parcela.status = 'EMITIDA'
        parcela.data_pagamento = None
    else:
        parcela.status = 'QUITADA'
        parcela.data_pagamento = date.today()
    parcela.save()
    return render(request, 'casos/partials/parcela_linha.html', {'parcela': parcela})

@login_required
@require_POST
def pagar_parcela(request, pk):
    parcela = get_object_or_404(Parcela, pk=pk)
    if parcela.status != 'QUITADA':
        parcela.status = 'QUITADA'
        parcela.data_pagamento = date.today()
        parcela.save()
        messages.success(request, 'Parcela quitada com sucesso.')
    return redirect(f"{reverse('casos:detalhe_caso', kwargs={'pk': parcela.acordo.caso.pk})}?aba=acordos")


@login_required
@require_POST
def upload_comprovante_parcela(request, pk):
    parcela = get_object_or_404(Parcela, pk=pk)
    arquivo = request.FILES.get('comprovante')
    if arquivo:
        parcela.comprovante = arquivo
        parcela.save()
        messages.success(request, 'Comprovante anexado com sucesso.')
    else:
        messages.error(request, 'Selecione um arquivo para anexar.')
    return redirect(f"{reverse('casos:detalhe_caso', kwargs={'pk': parcela.acordo.caso.pk})}?aba=acordos")


@login_required
def baixar_comprovante_parcela(request, pk):
    parcela = get_object_or_404(Parcela, pk=pk)
    if not parcela.comprovante:
        messages.error(request, 'Nenhum comprovante encontrado.')
        return redirect(f"{reverse('casos:detalhe_caso', kwargs={'pk': parcela.acordo.caso.pk})}?aba=acordos")
    response = FileResponse(parcela.comprovante.open('rb'), as_attachment=False)
    response['Content-Type'] = 'application/octet-stream'
    return response



@login_required

@login_required
@require_POST
def deletar_acordo(request, pk):
    acordo = get_object_or_404(Acordo, pk=pk)
    caso_pk = acordo.caso.pk
    acordo.delete()
    messages.success(request, 'Acordo excluido com sucesso.')
    return redirect(f"{reverse('casos:detalhe_caso', kwargs={'pk': caso_pk})}?aba=acordos")


@login_required
def exportar_acordo_excel(request, pk):
    acordo = get_object_or_404(Acordo, pk=pk)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Acordo_{acordo.pk}"
    headers = [
        'Acordo ID', 'Caso ID', 'Valor Total', 'Numero Parcelas',
        'Data Primeira Parcela', 'Advogado Responsavel', 'Data Criacao',
        'Parcela', 'Vencimento', 'Pagamento', 'Valor Parcela', 'Status'
    ]
    ws.append(headers)

    advogado = '-'
    if acordo.advogado_acordo:
        advogado = acordo.advogado_acordo.get_full_name() or acordo.advogado_acordo.username

    for parcela in acordo.parcelas.all().order_by('numero_parcela'):
        ws.append([
            acordo.pk,
            acordo.caso.pk,
            str(acordo.valor_total),
            acordo.numero_parcelas,
            acordo.data_primeira_parcela.strftime('%d/%m/%Y') if acordo.data_primeira_parcela else '-',
            advogado,
            acordo.data_criacao.strftime('%d/%m/%Y') if acordo.data_criacao else '-',
            parcela.numero_parcela,
            parcela.data_vencimento.strftime('%d/%m/%Y') if parcela.data_vencimento else '-',
            parcela.data_pagamento.strftime('%d/%m/%Y') if parcela.data_pagamento else '-',
            str(parcela.valor_parcela),
            parcela.get_status_display(),
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="acordo_{acordo.pk}.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_acordo_pdf(request, pk):
    acordo = get_object_or_404(Acordo, pk=pk)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="acordo_{acordo.pk}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()

    advogado = '-'
    if acordo.advogado_acordo:
        advogado = acordo.advogado_acordo.get_full_name() or acordo.advogado_acordo.username

    elements.append(Paragraph(f"Acordo #{acordo.pk} - Caso #{acordo.caso.pk}", styles['Title']))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Valor Total: R$ {acordo.valor_total}", styles['Normal']))
    elements.append(Paragraph(f"Numero de Parcelas: {acordo.numero_parcelas}", styles['Normal']))
    elements.append(Paragraph(f"Data Primeira Parcela: {acordo.data_primeira_parcela.strftime('%d/%m/%Y') if acordo.data_primeira_parcela else '-'}", styles['Normal']))
    elements.append(Paragraph(f"Advogado Responsavel: {advogado}", styles['Normal']))
    elements.append(Spacer(1, 12))

    data = [['Parcela', 'Vencimento', 'Pagamento', 'Valor', 'Status']]
    for parcela in acordo.parcelas.all().order_by('numero_parcela'):
        data.append([
            str(parcela.numero_parcela),
            parcela.data_vencimento.strftime('%d/%m/%Y') if parcela.data_vencimento else '-',
            parcela.data_pagamento.strftime('%d/%m/%Y') if parcela.data_pagamento else '-',
            f"R$ {parcela.valor_parcela}",
            parcela.get_status_display(),
        ])

    table = Table(data, colWidths=[60, 100, 100, 90, 90])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.orange),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))

    elements.append(table)
    doc.build(elements)
    return response


def editar_acordo(request, pk):
    acordo = get_object_or_404(Acordo, pk=pk)
    caso = acordo.caso
    if request.method == 'POST':
        form = AcordoForm(request.POST, instance=acordo, user=request.user)
        if form.is_valid():
            acordo_editado = form.save()
            acordo_editado.parcelas.all().delete()
            valor_total = acordo_editado.valor_total
            num_parcelas = acordo_editado.numero_parcelas
            valor_parcela = round(Decimal(valor_total) / num_parcelas, 2)
            for i in range(num_parcelas):
                data_vencimento = acordo_editado.data_primeira_parcela + relativedelta(months=i)
                Parcela.objects.create(
                    acordo=acordo_editado, numero_parcela=i + 1, valor_parcela=valor_parcela, data_vencimento=data_vencimento
                )
            soma_parcelas = valor_parcela * num_parcelas
            diferenca = valor_total - soma_parcelas
            if diferenca != 0:
                ultima_parcela = acordo_editado.parcelas.order_by('-numero_parcela').first()
                if ultima_parcela:
                    ultima_parcela.valor_parcela += diferenca
                    ultima_parcela.save()
            return redirect(f"{reverse('casos:detalhe_caso', kwargs={'pk': caso.pk})}?aba=acordos")
    else:
        form = AcordoForm(instance=acordo, user=request.user)
    context = {'form_acordo': form, 'acordo': acordo, 'caso': caso}
    return render(request, 'casos/acordo_form.html', context)


@login_required

@login_required
@require_POST
def deletar_despesa(request, pk):
    despesa = get_object_or_404(Despesa, pk=pk)
    caso_pk = despesa.caso.pk
    despesa.delete()
    messages.success(request, 'Despesa excluida com sucesso.')
    return redirect(f"{reverse('casos:detalhe_caso', kwargs={'pk': caso_pk})}?aba=despesas")


@login_required
@require_POST
def upload_comprovante_despesa(request, pk):
    despesa = get_object_or_404(Despesa, pk=pk)
    arquivo = request.FILES.get('comprovante')
    if arquivo:
        despesa.comprovante = arquivo
        despesa.save()
        messages.success(request, 'Comprovante anexado com sucesso.')
    else:
        messages.error(request, 'Selecione um arquivo para anexar.')
    return redirect(f"{reverse('casos:detalhe_caso', kwargs={'pk': despesa.caso.pk})}?aba=despesas")


@login_required
def baixar_comprovante_despesa(request, pk):
    despesa = get_object_or_404(Despesa, pk=pk)
    if not despesa.comprovante:
        messages.error(request, 'Nenhum comprovante encontrado.')
        return redirect(f"{reverse('casos:detalhe_caso', kwargs={'pk': despesa.caso.pk})}?aba=despesas")
    response = FileResponse(despesa.comprovante.open('rb'), as_attachment=False)
    response['Content-Type'] = 'application/octet-stream'
    return response


@login_required
def exportar_despesas_excel(request, pk):
    caso = get_object_or_404(Caso, pk=pk)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Despesas_Caso_{caso.pk}"
    headers = ['Caso ID', 'Data', 'Advogado', 'Descricao', 'Valor', 'Comprovante']
    ws.append(headers)

    for d in caso.despesas.all().order_by('-data_despesa'):
        advogado = '-'
        if d.advogado:
            advogado = d.advogado.get_full_name() or d.advogado.username
        ws.append([
            caso.pk,
            d.data_despesa.strftime('%d/%m/%Y') if d.data_despesa else '-',
            advogado,
            d.descricao,
            str(d.valor),
            d.comprovante.name if d.comprovante else '-'
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="despesas_caso_{caso.pk}.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_despesas_pdf(request, pk):
    caso = get_object_or_404(Caso, pk=pk)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="despesas_caso_{caso.pk}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph(f"Despesas - Caso #{caso.pk}", styles['Title']))
    elements.append(Spacer(1, 12))

    data = [['Data', 'Advogado', 'Descricao', 'Valor', 'Comprovante']]
    for d in caso.despesas.all().order_by('-data_despesa'):
        advogado = '-'
        if d.advogado:
            advogado = d.advogado.get_full_name() or d.advogado.username
        data.append([
            d.data_despesa.strftime('%d/%m/%Y') if d.data_despesa else '-',
            advogado,
            d.descricao,
            f"R$ {d.valor}",
            d.comprovante.name if d.comprovante else '-'
        ])

    table = Table(data, colWidths=[90, 120, 240, 80, 140])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.orange),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))

    elements.append(table)
    doc.build(elements)
    return response


def editar_despesa(request, pk):
    despesa = get_object_or_404(Despesa, pk=pk)
    caso = despesa.caso
    if request.method == 'POST':
        form = DespesaForm(request.POST, instance=despesa, user=request.user)
        if form.is_valid():
            form.save()
            return redirect(f"{reverse('casos:detalhe_caso', kwargs={'pk': caso.pk})}?aba=despesas")
    else:
        form = DespesaForm(instance=despesa, user=request.user)
    context = {'form_despesa': form, 'despesa': despesa, 'caso': caso}
    return render(request, 'casos/despesa_form.html', context)


# ==============================================================================
# API (DRF)
# ==============================================================================

class CasoAPIViewSet(viewsets.ModelViewSet):
    queryset = Caso.objects.all().order_by('-data_criacao')
    serializer_class = CasoSerializer
    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]

    def perform_update(self, serializer):
        if 'status' in serializer.validated_data:
            if serializer.validated_data['status'] == 'ENCERRADO':
                if not serializer.instance.data_encerramento:
                    serializer.validated_data['data_encerramento'] = timezone.now().date()
        serializer.save()


# ==============================================================================
# IMPORTAÇÃO/EXPORTAÇÃO DINÂMICA
# ==============================================================================

@login_required
def selecionar_filtros_exportacao(request):
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente')
        produto_id = request.POST.get('produto')
        if cliente_id and produto_id:
            return redirect('casos:exportar_casos_dinamico', cliente_id=cliente_id, produto_id=produto_id)
    clientes = Cliente.objects.all().order_by('nome')
    produtos = Produto.objects.all().order_by('nome')
    context = {'clientes': clientes, 'produtos': produtos, 'titulo': 'Exportação Dinâmica de Casos'}
    return render(request, 'casos/selecionar_filtros_exportacao.html', context)

@login_required
def exportar_casos_excel(request):
    casos_queryset = Caso.objects.select_related(
        'cliente', 'produto', 'advogado_responsavel'
    ).order_by('-data_entrada')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Casos'
    headers = [
        'ID', 'Cliente', 'Produto', 'Status', 'Data Entrada',
        'Advogado', 'Valor Apurado'
    ]
    sheet.append(headers)

    for caso in casos_queryset:
        advogado = ''
        if caso.advogado_responsavel:
            advogado = caso.advogado_responsavel.get_full_name() or caso.advogado_responsavel.username
        data_entrada = caso.data_entrada.strftime('%d/%m/%Y') if caso.data_entrada else ''
        sheet.append([
            caso.id,
            caso.cliente.nome if caso.cliente else '',
            caso.produto.nome if caso.produto else '',
            caso.get_status_display(),
            data_entrada,
            advogado,
            str(caso.valor_apurado) if caso.valor_apurado is not None else ''
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="casos.xlsx"'
    workbook.save(response)
    return response


@login_required
def exportar_casos_dinamico(request, cliente_id, produto_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    produto = get_object_or_404(Produto, id=produto_id)
    lista_chaves, lista_cabecalhos = get_cabecalho_exportacao(cliente, produto)
    casos_queryset = Caso.objects.filter(cliente=cliente, produto=produto).select_related(
        'cliente', 'produto', 'advogado_responsavel'
    ).prefetch_related('valores_personalizados__campo').order_by('-data_entrada')
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f'Export_{produto.nome[:30]}'
    sheet.append(lista_cabecalhos)
    sheet.append(lista_chaves)
    
    for caso in casos_queryset:
        linha_dados = []
        valores_personalizados_case = {f'personalizado_{v.campo.nome_variavel}': v.valor for v in caso.valores_personalizados.all()}
        for chave in lista_chaves:
            valor = '-'
            if not chave.startswith('personalizado_'):
                if '__' in chave:
                    partes = chave.split('__')
                    obj = getattr(caso, partes[0], None)
                    valor = getattr(obj, partes[1], '-') if obj else '-'
                elif chave == 'status':
                    valor = caso.get_status_display()
                else:
                    valor = getattr(caso, chave, '-')
            else:
                valor = valores_personalizados_case.get(chave, '-')
            linha_dados.append(str(valor))
        sheet.append(linha_dados)
        
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'casos_{cliente.nome}_{produto.nome}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def importar_casos_view(request):
    if request.method == 'GET':
        try:
            clientes = Cliente.objects.all().order_by('nome')
            produtos = Produto.objects.all().order_by('nome')
            context = {'clientes': clientes, 'produtos': produtos, 'titulo': 'Importação Massiva de Casos'}
            return render(request, 'casos/importar_casos_form.html', context)
        except Exception as e:
            logger.error(f"Erro ao carregar importação: {e}", exc_info=True)
            messages.error(request, "Erro ao carregar página.")
            return redirect('casos:lista_casos')

    elif request.method == 'POST':
        cliente_id = request.POST.get('cliente')
        produto_id = request.POST.get('produto')
        arquivo_excel = request.FILES.get('arquivo_excel')

        if not (cliente_id and produto_id and arquivo_excel):
            messages.error(request, "Todos os campos são obrigatórios.")
            return redirect('casos:importar_casos_view')

        try:
            cliente = get_object_or_404(Cliente, id=cliente_id)
            produto = get_object_or_404(Produto, id=produto_id)
            lista_chaves_validas, _ = get_cabecalho_exportacao(cliente, produto)
            chaves_validas_set = set(lista_chaves_validas)
            estrutura_campos = EstruturaDeCampos.objects.filter(cliente=cliente, produto=produto).prefetch_related('campos').first()
            campos_meta_map = {cm.nome_variavel: cm for cm in estrutura_campos.campos.all()} if estrutura_campos else {}
            
            workbook = openpyxl.load_workbook(arquivo_excel, data_only=True)
            sheet = workbook.active
            if sheet.max_row < 2:
                raise ValidationError("Planilha vazia.")

            excel_headers_raw = [cell.value for cell in sheet[1]]
            excel_headers = [str(h).strip().lower().replace(' ', '_') if h else '' for h in excel_headers_raw]
            header_map = {}
            variaveis_lower = {nome_var.lower(): nome_var for nome_var in campos_meta_map.keys()}

            for excel_header_norm in excel_headers:
                if not excel_header_norm: continue
                chave_mapeada = None
                if excel_header_norm in chaves_validas_set and not excel_header_norm.startswith('personalizado_') and '__' not in excel_header_norm:
                    chave_mapeada = excel_header_norm
                elif excel_header_norm in variaveis_lower:
                    nome_variavel_original = variaveis_lower[excel_header_norm]
                    if f'personalizado_{nome_variavel_original}' in chaves_validas_set:
                        chave_mapeada = nome_variavel_original
                elif excel_header_norm.startswith('personalizado_'):
                    nome_base = excel_header_norm.split('personalizado_', 1)[1]
                    if nome_base in variaveis_lower:
                        nome_variavel_original = variaveis_lower[nome_base]
                        if f'personalizado_{nome_variavel_original}' in chaves_validas_set:
                            chave_mapeada = nome_variavel_original
                if chave_mapeada:
                    header_map[excel_header_norm] = chave_mapeada

            mapeamentos_uteis = {k: v for k, v in header_map.items() if k != '_row_index'}
            if not mapeamentos_uteis:
                raise ValidationError("Nenhum cabeçalho corresponde aos campos esperados.")

            linhas_enviadas = 0
            campos_meta_map_serializable = {nome_var: campo.id for nome_var, campo in campos_meta_map.items()}

            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                linha_dados = dict(zip(excel_headers, row))
                linha_dados['_row_index'] = row_index
                linha_dados_mapeada = {k: v for k, v in linha_dados.items() if k in header_map or k == '_row_index'}
                if not any(v for k, v in linha_dados_mapeada.items() if k != '_row_index' and v is not None):
                    continue

                processar_linha_importacao.apply_async(
                    args=[
                        linha_dados_mapeada, cliente.id, produto.id, header_map, list(chaves_validas_set),
                        campos_meta_map_serializable, produto.padrao_titulo, estrutura_campos.id if estrutura_campos else None
                    ],
                    countdown=linhas_enviadas * 10
                )
                linhas_enviadas += 1
            
            if linhas_enviadas == 0:
                messages.warning(request, "Nenhuma linha válida encontrada.")
            else:
                messages.success(request, f"✅ Importação iniciada! {linhas_enviadas} casos enviados.")
            return redirect('casos:importar_casos_view')

        except ValidationError as e:
            messages.error(request, f"❌ Erro: {e.message}")
            return render(request, 'casos/importar_casos_form.html', {'clientes': Cliente.objects.all(), 'produtos': Produto.objects.all()})
        except Exception as e:
            logger.error(f"Erro inesperado: {e}", exc_info=True)
            messages.error(request, "❌ Erro inesperado.")
            return redirect('casos:importar_casos_view')

    return redirect('casos:importar_casos_view')

@login_required
@require_POST
def editar_info_basicas(request, pk):
    caso = get_object_or_404(Caso, pk=pk)
    caso.status = request.POST.get('status')
    caso.data_entrada = request.POST.get('data_entrada')
    valor_apurado = normalize_currency_input(request.POST.get('valor_apurado', ''))
    caso.valor_apurado = Decimal(valor_apurado) if valor_apurado else None
    caso.advogado_responsavel_id = request.POST.get('advogado_responsavel')
    caso.save()
    return render(request, 'casos/partials/card_info_basicas.html', {'caso': caso})

@login_required
@require_POST
def editar_dados_adicionais(request, pk):
    caso = get_object_or_404(Caso, pk=pk)
    caso.sinistro_todo = request.POST.get('sinistro_todo')
    caso.acao = request.POST.get('acao')
    caso.save()
    return render(request, 'casos/partials/card_dados_adicionais.html', {'caso': caso})

@login_required
def visao_casos_prazo(request):
    casos_list = Caso.objects.select_related('cliente', 'produto', 'advogado_responsavel').filter(status='ATIVO')
    filtro_cliente = request.GET.get('filtro_cliente', '')
    filtro_produto = request.GET.get('filtro_produto', '')
    filtro_advogado = request.GET.get('filtro_advogado', '')

    if filtro_cliente: casos_list = casos_list.filter(cliente_id=filtro_cliente)
    if filtro_produto: casos_list = casos_list.filter(produto_id=filtro_produto)
    if filtro_advogado: casos_list = casos_list.filter(advogado_responsavel_id=filtro_advogado)

    prazo_inicio_str = request.GET.get('prazo_inicio', '')
    prazo_fim_str = request.GET.get('prazo_fim', '')
    
    if prazo_inicio_str or prazo_fim_str:
        try:
            prazo_inicio = datetime.strptime(prazo_inicio_str, '%Y-%m-%d').date() if prazo_inicio_str else None
            prazo_fim = datetime.strptime(prazo_fim_str, '%Y-%m-%d').date() if prazo_fim_str else None
            casos_filtrados = []
            for caso in casos_list:
                prazo_final = caso.prazo_final_calculado
                if prazo_final:
                    if prazo_inicio and prazo_fim:
                        if prazo_inicio <= prazo_final <= prazo_fim: casos_filtrados.append(caso)
                    elif prazo_inicio:
                        if prazo_final >= prazo_inicio: casos_filtrados.append(caso)
                    elif prazo_fim:
                        if prazo_final <= prazo_fim: casos_filtrados.append(caso)
            casos_list = casos_filtrados
        except ValueError:
            messages.error(request, "Formato de data inválido.")
            casos_list = []
    
    casos_list = sorted(casos_list, key=lambda caso: caso.prazo_final_calculado or date.max)
    context = {
        'casos': casos_list,
        'valores_filtro': request.GET,
        'todos_clientes': Cliente.objects.all().order_by('nome'),
        'todos_produtos': Produto.objects.all().order_by('nome'),
        'todos_advogados': User.objects.filter(is_active=True).order_by('first_name'),
        'hoje': date.today(),
    }
    return render(request, 'casos/visao_casos_prazo.html', context)

@login_required
def analyser_navegador(request, pk):
    caso = get_object_or_404(Caso, pk=pk)
    if not caso.sharepoint_folder_id:
        return HttpResponse('<div class="analyser-empty-state">Pasta não encontrada</div>')
    try:
        sp = SharePoint()
        itens = sp.listar_conteudo_pasta(caso.sharepoint_folder_id)
        return render(request, 'casos/partials/painel_anexos.html', {
            'caso': caso, 'itens': itens, 'folder_id': caso.sharepoint_folder_id
        })
    except Exception as e:
        return HttpResponse(f"Erro SharePoint: {e}")

@login_required
def upload_arquivo_sharepoint(request, caso_pk):
    caso = get_object_or_404(Caso, pk=caso_pk)
    if request.method == 'POST' and request.FILES.get('arquivo'):
        sp = SharePoint()
        pasta_id = request.POST.get('pasta_id', caso.sharepoint_folder_id)
        sp.fazer_upload(request.FILES['arquivo'], pasta_id)
        return carregar_painel_anexos(request, caso_pk)
    return JsonResponse({'error': 'Falha no upload'}, status=400)

@require_POST
@login_required
def criar_pasta_raiz_sharepoint(request):
    try:
        nome_nova_pasta = request.POST.get('nome_pasta')
        if not nome_nova_pasta:
            return HttpResponse("<p style='color: red;'>Nome da pasta nao pode ser vazio.</p>", status=400)
        sp = SharePoint()
        sp.criar_pasta_caso(nome_nova_pasta)
    except Exception as e:
        logger.error(f"Erro ao criar pasta na raiz: {e}", exc_info=True)
        return HttpResponse(f"<p style='color: red;'>Erro ao criar pasta: {e}</p>", status=500)
    response = HttpResponse(status=200)
    response['HX-Refresh'] = 'true'
    return response

# ==============================================================================
# OUTRAS VIEWS (TIMESHEET, ACORDOS, ETC)
# ==============================================================================
@login_required
def editar_timesheet(request, pk):
    ts = get_object_or_404(Timesheet, pk=pk)
    if request.method == 'POST':
        form = TimesheetForm(request.POST, instance=ts)
        if form.is_valid():
            form.save()
            return redirect('casos:detalhe_caso', pk=ts.caso.pk)
    return render(request, 'casos/timesheet_form.html', {'form': TimesheetForm(instance=ts)})

@login_required
def selecionar_produto_cliente(request):
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente')
        produto_id = request.POST.get('produto')
        if cliente_id and produto_id:
            return redirect('casos:criar_caso', cliente_id=cliente_id, produto_id=produto_id)
        messages.error(request, "Selecione um cliente e um produto.")
    clientes = Cliente.objects.all().order_by('nome')
    produtos = Produto.objects.all().order_by('nome')
    return render(request, 'casos/selecionar_produto_cliente.html', {'clientes': clientes, 'produtos': produtos})

@login_required
def obter_detalhes_tomador(request, pk):
    """
    Busca os dados do tomador (CPF, Emails, Telefones) e retorna JSON.
    """
    try:
        tomador = Tomador.objects.get(pk=pk)
        
        # Pega a lista de emails e telefones
        emails = list(tomador.emails.values_list('email', flat=True))
        telefones = [
            f"{fone.get_tipo_display()}: {fone.telefone}"
            for fone in tomador.telefones.all()
        ]
        
        return JsonResponse({
            'success': True,
            'tipo': tomador.tipo,
            'cpf': tomador.cpf,
            'cnpj': tomador.cnpj,
            'documento': tomador.cpf if tomador.tipo == 'PF' else tomador.cnpj,
            'emails': emails,
            'telefones': telefones
        })
    except Tomador.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Tomador não encontrado'})


@login_required
def obter_detalhes_segurado(request, pk):
    """
    Busca os dados do segurado (CPF, Emails, Telefones) e retorna JSON.
    """
    try:
        segurado = Segurado.objects.get(pk=pk)
        emails = list(segurado.emails.values_list('email', flat=True))
        telefones = [
            f"{fone.get_tipo_display()}: {fone.telefone}"
            for fone in segurado.telefones.all()
        ]
        return JsonResponse({
            'success': True,
            'tipo': segurado.tipo,
            'cpf': segurado.cpf,
            'cnpj': segurado.cnpj,
            'documento': segurado.cpf if segurado.tipo == 'PF' else segurado.cnpj,
            'emails': emails,
            'telefones': telefones
        })
    except Segurado.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Segurado nao encontrado'})


@login_required
def obter_detalhes_corretor(request, pk):
    """
    Busca os dados do corretor (CPF, Emails, Telefones) e retorna JSON.
    """
    try:
        corretor = Corretor.objects.get(pk=pk)
        emails = list(corretor.emails.values_list('email', flat=True))
        telefones = [
            f"{fone.get_tipo_display()}: {fone.telefone}"
            for fone in corretor.telefones.all()
        ]
        return JsonResponse({
            'success': True,
            'tipo': corretor.tipo,
            'cpf': corretor.cpf,
            'cnpj': corretor.cnpj,
            'documento': corretor.cpf if corretor.tipo == 'PF' else corretor.cnpj,
            'emails': emails,
            'telefones': telefones
        })
    except Corretor.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Corretor nao encontrado'})
